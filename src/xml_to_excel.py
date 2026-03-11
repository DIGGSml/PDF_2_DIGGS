"""
xml_to_excel.py
Converts a DIGGS 2.6 XML borehole file into:
  1. A single-borehole Excel (2 sheets: Borehole Info + Stratigraphy & SPT)
  2. Upserts rows into a master multi-borehole Excel
"""

import os
import re
import logging
from lxml import etree
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ─── Namespace map (matches diggs_generator.py) ──────────────────────────────
NS = {
    "d":    "http://diggsml.org/schemas/2.6",
    "gml":  "http://www.opengis.net/gml/3.2",
    "xlink":"http://www.w3.org/1999/xlink",
}

# ─── Style helpers ────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1E3A5F")   # dark navy
SUBHEAD_FILL  = PatternFill("solid", fgColor="2E5090")   # medium blue
EVEN_FILL     = PatternFill("solid", fgColor="EBF0FA")   # light blue-white
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
SUBHEAD_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
LABEL_FONT    = Font(bold=True, name="Calibri", size=10)
DATA_FONT     = Font(name="Calibri", size=10)
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN    = Alignment(horizontal="left",  vertical="center", wrap_text=True)
THIN_BORDER   = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

SOIL_COLORS = {
    "SAND":  "F4A460",
    "SILT":  "D2B48C",
    "CLAY":  "A0522D",
    "ROCK":  "808080",
    "FILL":  "9E9E9E",
}


def _soil_fill(description):
    if not description:
        return WHITE_FILL
    upper = description.upper()
    for key, hex_color in SOIL_COLORS.items():
        if key in upper:
            return PatternFill("solid", fgColor=hex_color)
    return WHITE_FILL


def _cell(ws, row, col, value, font=None, fill=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = border
    return c


# ─── XML parsing ──────────────────────────────────────────────────────────────

def parse_diggs_xml(xml_path):
    """
    Parse a DIGGS XML file and return a dict:
    {
      "borehole_id": str,
      "project_name": str,
      "latitude": float,
      "longitude": float,
      "elevation_ft": float,
      "total_depth_ft": float | None,
      "drill_date": str | None,
      "layers": [{"top": float, "bottom": float, "description": str, "uscs": str, "n_value": int|None}, ...]
    }
    """
    tree  = etree.parse(xml_path)
    root  = tree.getroot()

    # ── Borehole meta ──
    bh_el        = root.find(".//d:Borehole", NS)
    borehole_id  = bh_el.findtext("gml:name", default="Unknown", namespaces=NS) if bh_el else "Unknown"

    project_el   = root.find(".//d:Project", NS)
    project_name = project_el.findtext("gml:name", default="Unknown", namespaces=NS) if project_el else "Unknown"

    pos_text   = root.findtext(".//d:PointLocation/gml:pos", default="0 0 0", namespaces=NS)
    pos_parts  = pos_text.strip().split()
    try:
        lat  = float(pos_parts[0]) if len(pos_parts) > 0 else None
        lon  = float(pos_parts[1]) if len(pos_parts) > 1 else None
        elev = float(pos_parts[2]) if len(pos_parts) > 2 else 0.0
    except (ValueError, IndexError):
        lat, lon, elev = None, None, 0.0

    depth_el    = root.find(".//d:totalMeasuredDepth", NS)
    total_depth = float(depth_el.text) if depth_el is not None and depth_el.text else None

    date_el    = root.find(".//d:TimeInterval/d:end", NS)
    drill_date = None
    if date_el is not None and date_el.text:
        drill_date = date_el.text.split("T")[0]   # strip time component

    # ── Lithology layers ──
    lith_map = {}   # (top, bottom) -> {"description": str, "uscs": str}
    for obs_el in root.findall(".//d:LithologySystem", NS):
        pos_list = obs_el.findtext(".//d:LinearExtent/gml:posList", default="", namespaces=NS)
        try:
            top_d, bot_d = [float(x) for x in pos_list.strip().split()]
        except ValueError:
            continue
        desc = obs_el.findtext(".//d:lithDescription", default="", namespaces=NS)
        uscs = obs_el.findtext(".//d:legendCode",      default="", namespaces=NS)
        lith_map[(top_d, bot_d)] = {"description": desc.strip(), "uscs": uscs.strip()}

    # ── SPT results ──
    spt_map = {}    # (top, bottom) -> n_value
    for test_el in root.findall(".//d:Test", NS):
        name_el = test_el.find("gml:name", NS)
        if name_el is None or "SPT" not in (name_el.text or ""):
            continue
        pos_list = test_el.findtext(".//d:LinearExtent/gml:posList", default="", namespaces=NS)
        try:
            top_d, bot_d = [float(x) for x in pos_list.strip().split()]
        except ValueError:
            continue
        val_text = test_el.findtext(".//d:dataValues", default="", namespaces=NS)
        try:
            n_val = int(float(val_text.strip()))
        except (ValueError, AttributeError):
            n_val = None
        spt_map[(top_d, bot_d)] = n_val

    # ── Merge layers ──
    all_depth_pairs = sorted(set(list(lith_map.keys()) + list(spt_map.keys())))
    layers = []
    for (top_d, bot_d) in all_depth_pairs:
        lith_info = lith_map.get((top_d, bot_d), {})
        layers.append({
            "top":         top_d,
            "bottom":      bot_d,
            "description": lith_info.get("description", ""),
            "uscs":        lith_info.get("uscs", ""),
            "n_value":     spt_map.get((top_d, bot_d)),
        })

    return {
        "borehole_id":   borehole_id,
        "project_name":  project_name,
        "latitude":      lat,
        "longitude":     lon,
        "elevation_ft":  elev,
        "total_depth_ft": total_depth,
        "drill_date":    drill_date,
        "layers":        layers,
    }


# ─── Single-borehole Excel ────────────────────────────────────────────────────

def xml_to_excel(xml_path, output_xlsx_path):
    """
    Converts a single DIGGS XML file to a clean 2-sheet Excel workbook.
    Sheet 1: Borehole Info  (key-value metadata)
    Sheet 2: Stratigraphy & SPT  (tabular layer data)
    """
    data = parse_diggs_xml(xml_path)
    wb   = openpyxl.Workbook()

    # ── Sheet 1: Borehole Info ──
    ws1 = wb.active
    ws1.title = "Borehole Info"
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 40

    # Title row
    ws1.merge_cells("A1:B1")
    _cell(ws1, 1, 1, "BOREHOLE SUMMARY", font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN)
    ws1.row_dimensions[1].height = 28

    # Sub-header
    ws1.merge_cells("A2:B2")
    _cell(ws1, 2, 1, f"Project: {data['project_name']}", font=SUBHEAD_FONT, fill=SUBHEAD_FILL, align=LEFT_ALIGN)
    ws1.row_dimensions[2].height = 22

    meta_rows = [
        ("Borehole ID",     data["borehole_id"]),
        ("Latitude",        data["latitude"]),
        ("Longitude",       data["longitude"]),
        ("Elevation (ft)",  data["elevation_ft"]),
        ("Total Depth (ft)",data["total_depth_ft"]),
        ("Drill Date",      data["drill_date"] or "N/A"),
    ]
    for i, (label, value) in enumerate(meta_rows):
        r = i + 3
        fill = EVEN_FILL if i % 2 == 0 else WHITE_FILL
        _cell(ws1, r, 1, label, font=LABEL_FONT, fill=fill, align=LEFT_ALIGN, border=THIN_BORDER)
        _cell(ws1, r, 2, value, font=DATA_FONT,  fill=fill, align=LEFT_ALIGN, border=THIN_BORDER)
        ws1.row_dimensions[r].height = 20

    # ── Sheet 2: Stratigraphy & SPT ──
    ws2 = wb.create_sheet("Stratigraphy & SPT")
    headers = ["Top Depth (ft)", "Bottom Depth (ft)", "Soil Description", "USCS Code", "N-Value (SPT)"]
    col_widths = [16, 18, 45, 12, 14]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        ws2.column_dimensions[get_column_letter(col)].width = w
        c = _cell(ws2, 1, col, h, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER_ALIGN, border=THIN_BORDER)

    ws2.row_dimensions[1].height = 26

    for i, layer in enumerate(data["layers"]):
        r = i + 2
        soil_fill = _soil_fill(layer["description"])
        row_data  = [layer["top"], layer["bottom"], layer["description"], layer["uscs"], layer["n_value"]]
        for col, val in enumerate(row_data, start=1):
            fill = soil_fill if col == 3 else (EVEN_FILL if i % 2 == 0 else WHITE_FILL)
            _cell(ws2, r, col, val, font=DATA_FONT, fill=fill, align=LEFT_ALIGN, border=THIN_BORDER)
        ws2.row_dimensions[r].height = 18

    # Freeze top row
    ws2.freeze_panes = "A2"

    wb.save(output_xlsx_path)
    logger.info(f"Single-borehole Excel saved → {output_xlsx_path}")
    return output_xlsx_path


# ─── Master multi-borehole Excel (upsert) ─────────────────────────────────────

MASTER_COLUMNS = [
    "Borehole ID", "Project Name", "Latitude", "Longitude",
    "Elevation (ft)", "Total Depth (ft)", "Drill Date",
    "Top Depth (ft)", "Bottom Depth (ft)", "Soil Description", "USCS Code", "N-Value (SPT)"
]

def upsert_master_excel(xml_path, master_path):
    """
    Parses the XML and upserts its borehole data into a master flat-table Excel.
    - If master_path doesn't exist: creates it with headers.
    - If borehole already in master: replaces all its rows.
    - If new borehole: appends rows at end.
    Returns the path to the master Excel.
    """
    data        = parse_diggs_xml(xml_path)
    borehole_id = data["borehole_id"]

    # Build rows to insert for this borehole
    new_rows = []
    for layer in data["layers"]:
        new_rows.append([
            borehole_id,
            data["project_name"],
            data["latitude"],
            data["longitude"],
            data["elevation_ft"],
            data["total_depth_ft"],
            data["drill_date"],
            layer["top"],
            layer["bottom"],
            layer["description"],
            layer["uscs"],
            layer["n_value"],
        ])
    # If no layers, still write a single metadata row
    if not new_rows:
        new_rows.append([
            borehole_id, data["project_name"], data["latitude"], data["longitude"],
            data["elevation_ft"], data["total_depth_ft"], data["drill_date"],
            None, None, None, None, None,
        ])

    if os.path.exists(master_path):
        wb  = openpyxl.load_workbook(master_path)
        ws  = wb.active
        # Read existing data, strip rows belonging to this borehole
        existing_rows = []
        header_row    = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header_row = list(row)
                continue
            if row[0] != borehole_id:   # col 0 = Borehole ID
                existing_rows.append(list(row))

        # Rebuild sheet
        ws.delete_rows(1, ws.max_row)
        _write_master_sheet(ws, header_row or MASTER_COLUMNS, existing_rows + new_rows)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "All Boreholes"
        _write_master_sheet(ws, MASTER_COLUMNS, new_rows)

    wb.save(master_path)
    logger.info(f"Master Excel upserted (borehole={borehole_id}) → {master_path}")
    return master_path


def _write_master_sheet(ws, headers, data_rows):
    """Write headers + data rows with consistent styling to a worksheet."""
    col_widths = [14, 28, 11, 12, 14, 15, 12, 14, 16, 42, 11, 14]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
        _cell(ws, 1, col, h, font=HEADER_FONT, fill=HEADER_FILL,
              align=CENTER_ALIGN, border=THIN_BORDER)
    ws.row_dimensions[1].height = 26

    for i, row in enumerate(data_rows):
        r         = i + 2
        soil_desc = row[9] if len(row) > 9 else ""
        for col, val in enumerate(row, start=1):
            fill = _soil_fill(soil_desc) if col == 10 else (EVEN_FILL if i % 2 == 0 else WHITE_FILL)
            _cell(ws, r, col, val, font=DATA_FONT, fill=fill,
                  align=LEFT_ALIGN, border=THIN_BORDER)
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A2"
